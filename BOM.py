import csv
import win32com.client as win32
import openpyxl

output_file_root = str()
import_file_root = str()
strip_parameters_file_root = str('C:\\Users\\PDAL\\Desktop\\VB-GTCA022\\strip_parameter.csv')
Mode_status = str('閉模')
input_root = str('C:\\Users\\PDAL\\Desktop\\VB-GTCA022\\auto\\catia_input-GTCA022\\')
# 檔案路徑
file_path = str('C:\\Users\\PDAL\\Desktop\\VB-GTCA022')
# 儲存路徑 (output 零件)
save_path = str(file_path + '\\auto\\catia_output-GTCA022\\')
# 母檔輸入路徑 (input Data)
open_path = str(file_path + "\\auto\\catia_input-GTCA022\\")
# 模具規範路徑
die_rule_path = str(file_path + "\\auto\\die_rule\\")
# 2D出圖路徑
drafting_output_path = str(file_path + "\\auto\\drafting_output-GTCA022\\")
# 標準零件路徑
standard_path = str(file_path + "\\auto\\Standard_Assembly\\")
# 製作一半的BOM表儲存路徑
onwork_BOM_open = str(file_path + "\\BOM表\\")
# BOM表儲存路徑
BOM_output_path = str(file_path + "\\auto\\BOM_output-GTCA022\\")
serch_result = float()
all_part_name = ['']

with open(strip_parameters_file_root) as csvFile:
    rows = csv.reader(csvFile)
    parameter_list = tuple(tuple(rows)[0])
    strip_parameter_list = parameter_list

Sheets = [0] * 99


def BOMMaking():
    create_catia_bom()
    (page) = output_bom()
    information_bom(page)
    save()


def create_catia_bom():
    catapp = win32.Dispatch("CATIA.Application")
    document = catapp.ActiveDocument
    product1 = document.Product
    assemblyConvertor1 = product1.getItem("BillOfMaterial")

    assemblyConvertor1Variant = assemblyConvertor1
    # assemblyConvertor1Variant.SetSecondaryFormat(arrayOfVariantOfBSTR1)
    arrayOfVariantOfBSTR1 = ["Quantity", "Part Number", "Material_Data", "Heat Treatment", "Product Description",
                             "Page", "Size"]
    assemblyConvertor1Variant.SetCurrentFormat(arrayOfVariantOfBSTR1)

    # 含數據內容之BOM表(複製用)儲存路徑
    assemblyConvertor1.Print("XLS", "" + BOM_output_path + "catia_bom.xlsx", product1)


def output_bom():
    wb1 = openpyxl.load_workbook(str(BOM_output_path) + "catia_bom.xlsx")
    wb2 = openpyxl.load_workbook(str(onwork_BOM_open) + "BOM_空白頁.xlsx")
    (cunt) = decide_Row()  # 搜尋資料數目
    (page) = decide_Page(cunt)  # page
    decide_Size(cunt, page)  # 規格
    decide_NO(cunt, page)  # 件號
    decide_name(cunt, page)  # 名稱
    decide_Quantity(cunt, page)  # 數量
    decide_material(cunt, page)  # 材質
    decide_Heat_treatment(cunt, page)  # 熱處理
    decide_description(cunt, page)  # 規格
    decide_Pa(cunt, page)
    # decide_cost(cunt, page)  # 價格
    draw_block(cunt, page)  # 備註
    # xlApp = GetObject("Excel.Application")  # 引用正在執行的Excel Application
    # xlBook = xlApp.openpyxl.load_workbook(str((onwork_BOM_open) + "BOM_空白頁.xlsx"))  # 打開已經存在的EXCEL工件簿文件
    # xlApp.Visible = True  # 要開啟，不然程式會隱藏

    return page


def information_bom(page):
    wb1 = openpyxl.load_workbook(str(str(BOM_output_path) + "BOM_空白頁.xlsx"))
    for j in range(1, page + 1):
        Sheetname = str(str("Sheet") + str(j))
        Sheets(Sheetname).Select()
        data = ["製    表", "製表日期", "頁    數", "模具編號", "品    號", "品    名"]
        kss = {"What": "製    表", "After": "ActiveCell", "LookIn": "xlFormulas", "LookAt": "xlPart",
               "SearchOrder": "xlByRows", "SearchDirection": "xlNext", "MatchCase": False, "MatchByte": False,
               "SearchFormat": False}
        for Data in data:
            if Data == "製    表":
                ss = kss.Column  # 列數
                sc = kss.Row
                Cells(sc, ss).Offset(ColumnOffset=1).Select()
                ActiveCell.FormulaR1C1 = Frame_Commissioner
            elif Data == "製表日期":
                kss[What] = "製表日期"
                ss = kss.Column  # 列數
                sc = kss.Row
                Cells(sc, ss).Offset(ColumnOffset=1).Select()
                ActiveCell.FormulaR1C1 = Frame_Todate
            elif Data == "頁    數":
                kss[What] = "頁    數"
                ss = kss.Column  # 列數
                sc = kss.Row
                Cells(sc, ss).Offset(ColumnOffset=1).Select()
                ActiveCell.FormulaR1C1 = j
            elif Data == "模具編號":
                kss[What] = "模具編號"
                ss = kss.Column  # 列數
                sc = kss.Row
                Cells(sc, ss).Offset(ColumnOffset=1).Select()
                ActiveCell.FormulaR1C1 = Frame_guest_number
            elif Data == "品    號":
                kss[What] = "品    號"
                ss = kss.Column  # 列數
                sc = kss.Row
                Cells(sc, ss).Offset(ColumnOffset=1).Select()
                ActiveCell.FormulaR1C1 = Frame_guest_number
            elif Data == "品    名":
                kss[What] = "品    名"
                ss = kss.Column  # 列數
                sc = kss.Row
                Cells(sc, ss).Offset(ColumnOffset=1).Select()
                ActiveCell.FormulaR1C1 = Finished_product_Name
        ActiveSheet.Range("A1").Select()
        ActiveCell.FormulaR1C1 = Company_Name
    Adjustment()


def save():
    wb = openpyxl.Workbook(BOM_output_path + "catia_bom.xlsx")
    write_BOM_location = str(BOM_output_path) + "BOM.xls"  # 最後BOM表存檔
    wb.save(write_BOM_location)
    FileName = write_BOM_location
    FileFormat = xlNormal, Password = ""
    WriteResPassword = ""
    ReadOnlyRecommended = False
    CreateBackup = False

    # ActiveWorkbook.SaveAs()
    # FileName = str("C:\\Documents and Settings\\SOA\\桌面\\BOM.xls")
    # FileFormat = xlNormal
    # Password = ""
    # WriteResPassword = ""
    # ReadOnlyRecommended = False
    # CreateBackup = False


def decide_Row():  # 判斷資料數目
    wb = openpyxl.load_workbook(BOM_output_path + "catia_bom.xlsx", data_only=False)

    Rng1 = {"what": "Quantity", "After": "ActiveCell", "LookIn": "xlFormulas",
            "LookAt": "xlPart", "SearchOrder": "xlByRows", "SearchDirection": "xlNext",
            "MatchCase": False, "MatchByte": False, "SearchFormat": False}

    sheet = wb['工作表1']

    for row in sheet.iter_rows(min_row=5, max_col=1, max_row=99, values_only=True):
        cunt = 0
        for value in row:
            if value == "":
                cunt -= 1
                break
            cunt += 1


def decide_Page(cunt):
    wb = openpyxl.load_workbook(str(str(BOM_output_path) + "BOM_空白頁.xlsx"))
    page = int(cunt / 30)
    if page < 1:
        page = 0
    for i in range(page, 1, -1):
        j = i
        sheet = wb['工作表1']
        target = wb.copy_worksheet(sheet)
        target.title = '工作表' + str(i)

    if page < 1:
        pagenumb = cunt
    if page >= 1:
        pagenumb = 30

    return page


def decide_Size(cunt, page):
    loops = 0
    if page < 1:
        pagenumb = cunt
    if page >= 1:
        pagenumb = 30
    page0 = page

    for j in range(1, page + 2):
        Sheetname = "Sheet" + str(j)
        ws =
    for i in range(1, pagenumb + 1):
        wb = openpyxl.load_workbook(str(str(BOM_output_path) + "catia_bom.xlsx"))
        kss = {"What": "Size", "After": "ActiveCell", "LookIn": "xlFormulas", "LookAt": "xlPart",
               "SearchOrder": "xlByRows", "SearchDirection": "xlNext", "MatchCase": False, "MatchByte": False,
               "SearchFormat": False}
        cells.Find(kss)
        ss = kss.Column  # 列數
        sc = kss.Row
        sc = kss.Row + loops  # 行動
        cells(sc, ss).Select()  # 位置
        cells(sc, ss).Offset(RowOffset=1).Select()
        Selection.Copy()
        wb = openpyxl.load_workbook(str(str(BOM_output_path) + "BOM_空白頁.xlsx"))
        kss1 = {"What": "規格", "After": "ActiveCell", "LookIn": "xlFormulas", "LookAt": "xlPart",
                "SearchOrder": "xlByRows", "SearchDirection": "xlNext", "MatchCase": False, "MatchByte": False,
                "SearchFormat": False}
        Cells.Find(kss1)
        ss1 = kss.Column  # 列數
        sc1 = kss.Row + i - 1  # 行動
        Cells(sc1, ss1).Select()  # 位置
        Cells(sc1, ss1).Offset(RowOffset=1).Select()
        ActiveSheet.Paste()
        loops += 1
    page0 -= 1
    if page0 == 0:
        pagenumb = cunt - 30 * page


def decide_NO(cunt, page):
    wb = openpyxl.load_workbook(str(str(BOM_output_path) + "BOM_空白頁.xlsx"))
    page0 = page
    for j in range(1, page + 1):
        Sheetname = str("Sheet" + str(j))
        Sheets(Sheetname).Select()
        for i in range(1, pagenumb + 1):
            kss = {"What": "件號", "After": "ActiveCell", "LookIn": "xlFormulas", "LookAt": "xlPart",
                   "SearchOrder": "xlByRows", "SearchDirection": "xlNext", "MatchCase": False, "MatchByte": False,
                   "SearchFormat": False}
            Cells.Find(kss)
            ss = kss.Column  # 列數
            sc = kss.Row + i - 1  # 行動
            Cells(sc, ss).Select()  # 位置
            Cells(sc, ss).Offset(RowOffset=1).Select()
            ActiveCell.FormulaR1C1 = i
        page0 -= 1
        if page0 == 0:
            pagenumb = cunt - 30 * page


def decide_name(cunt, page):
    loops = 0
    if page < 1:
        pagenumb = cunt
    if page >= 1:
        pagenumb = 30
    page0 = page
    for j in range(1, page + 1):
        Sheetname = str("Sheet" + str(j))
        Sheets(Sheetname).Select()
        for i in range(1, pagenumb):
            wb = openpyxl.load_workbook(str(str(BOM_output_path) + "catia_bom.xlsx"))
            kss = {"What": "Part Number", "After": "ActiveCell", "LookIn": "xlFormulas", "LookAt": "xlPart",
                   "SearchOrder": "xlByRows", "SearchDirection": "xlNext", "MatchCase": False, "MatchByte": False,
                   "SearchFormat": False}
            Cells.Find(kss)
            ss = kss.Column  # 列數
            sc = kss.Row + i - 1  # 行動
            Cells(sc, ss).Select()  # 位置
            Cells(sc, ss).Offset(RowOffset=1).Select()
            Selection.Copy()
            wb = openpyxl.load_workbook(str(str(BOM_output_path) + "BOM_空白頁.xlsx"))
            kss1 = {What: "名稱", After: ActiveCell, LookIn: xlFormulas, LookAt: xlPart, SearchOrder: xlByRows,
                    SearchDirection: xlNext, MatchCase: False, MatchByte: False, SearchFormat: False}
            Cells.Find(kss1)
            ss1 = kss.Column  # 列數
            sc1 = kss.Row + i - 1  # 行動
            Cells(sc1, ss1).Select()  # 位置
            Cells(sc1, ss1).Offset(RowOffset=1).Select()
            ActiveSheet.Paste()
            loops += 1
        page0 -= 1
        if page0 == 0:
            pagenumb = cunt - 30 * page


def decide_Quantity(cunt, page):
    loops = 0
    if page < 1:
        pagenumb = cunt
    if page >= 1:
        pagenumb = 30
    page0 = page
    for j in range(1, page + 1):
        Sheetname = str("Sheet" + str(j))
        Sheets(Sheetname).Select()
        for i in range(1, pagenumb):
            wb = openpyxl.load_workbook(str(str(BOM_output_path) + "catia_bom.xlsx"))
            kss = {"What": "Quantity", "After": "ActiveCell", "LookIn": "xlFormulas", "LookAt": "xlPart",
                   "SearchOrder": "xlByRows", "SearchDirection": "xlNext", "MatchCase": False, "MatchByte": False,
                   "SearchFormat": False}
            Cells.Find(kss)
            ss = kss.Column  # 列數
            sc = kss.Row + i - 1  # 行動
            Cells(sc, ss).Select()  # 位置
            Cells(sc, ss).Offset(RowOffset=1).Select()
            Selection.Copy()
            wb = openpyxl.load_workbook(str(str(BOM_output_path) + "BOM_空白頁.xlsx"))
            kss1 = {What: "數量", After: ActiveCell, LookIn: xlFormulas, LookAt: xlPart, SearchOrder: xlByRows,
                    SearchDirection: xlNext, MatchCase: False, MatchByte: False, SearchFormat: False}
            Cells.Find(kss1)
            ss1 = kss.Column  # 列數
            sc1 = kss.Row + i - 1  # 行動
            Cells(sc1, ss1).Select()  # 位置
            Cells(sc1, ss1).Offset(RowOffset=1).Select()
            ActiveSheet.Paste()
            loops += 1
        page0 -= 1
        if page0 == 0:
            pagenumb = cunt - 30 * page


def decide_material(cunt, page):
    loops = 0
    if page < 1:
        pagenumb = cunt
    if page >= 1:
        pagenumb = 30
    page0 = page
    for j in range(1, page + 1):
        Sheetname = str("Sheet" + str(j))
        Sheets(Sheetname).Select()
        for i in range(1, pagenumb + 1):
            wb = openpyxl.load_workbook(str(str(BOM_output_path) + "catia_bom.xlsx"))
            kss = {"What": "Material_Data", "After": "ActiveCell", "LookIn": "xlFormulas", "LookAt": "xlPart",
                   "SearchOrder": "xlByRows", "SearchDirection": "xlNext", "MatchCase": False, "MatchByte": False,
                   "SearchFormat": "False"}
            Cells.Find(kss)
            ss = kss.Column  # 列數
            sc = kss.Row + i - 1  # 行動
            Cells(sc, ss).Select()  # 位置
            Cells(sc, ss).Offset(RowOffset=1).Select()
            Selection.Copy()
            wb = openpyxl.load_workbook(str(str(BOM_output_path) + "BOM_空白頁.xlsx"))
            kss1 = {What: "材質", After: ActiveCell, LookIn: xlFormulas, LookAt: xlPart, SearchOrder: xlByRows,
                    SearchDirection: xlNext, MatchCase: False, MatchByte: False, SearchFormat: False}
            Cells.Find(kss1)
            ss1 = kss.Column  # 列數
            sc1 = kss.Row + i - 1  # 行動
            Cells(sc1, ss1).Select()  # 位置
            Cells(sc1, ss1).Offset(RowOffset=1).Select()
            ActiveSheet.Paste()
            loops += 1
            page0 -= 1
            if page0 == 0:
                pagenumb = cunt - 30 * page


def decide_Heat_treatment(cunt, page):
    loops = 0
    if page < 1:
        pagenumb = cunt
    if page >= 1:
        pagenumb = 30
    page0 = page
    for j in range(1, page + 1):
        Sheetname = str("Sheet" + str(j))
        Sheets(Sheetname).Select()
        for i in range(1, pagenumb):
            wb = openpyxl.load_workbook(str(str(BOM_output_path) + "catia_bom.xlsx"))
            kss = {"What": "Heat Treatment", "After": "ActiveCell", "LookIn": "xlFormulas", "LookAt": "xlPart",
                   "SearchOrder": "xlByRows", "SearchDirection": "xlNext", "MatchCase": False, "MatchByte": False,
                   "SearchFormat": False}
            Cells.Find(kss)
            ss = kss.Column  # 列數
            sc = kss.Row + i - 1  # 行動
            Cells(sc, ss).Select()  # 位置
            Cells(sc, ss).Offset(RowOffset=1).Select()
            Selection.Copy()
            wb = openpyxl.load_workbook(str(str(BOM_output_path) + "BOM_空白頁.xlsx"))
            kss1 = {"What": "熱處理", "After": "ActiveCell", "LookIn": "xlFormulas", "LookAt": "xlPart",
                    "SearchOrder": "xlByRows", "SearchDirection": "xlNext", "MatchCase": False, "MatchByte": False,
                    "SearchFormat": False}
            Cells.Find(kss1)
            ss1 = kss.Column  # 列數
            sc1 = kss.Row + i - 1  # 行動
            Cells(sc1, ss1).Select()  # 位置
            Cells(sc1, ss1).Offset(RowOffset=1).Select()
            ActiveSheet.Paste()
            loops += 1
        page0 -= 1
        if page0 == 0:
            pagenumb = cunt - 30 * page


def decide_description(cunt, page):
    loops = 0
    if page < 1:
        pagenumb = cunt
    if page >= 1:
        pagenumb = 30
    page0 = page
    for j in range(1, page + 1):
        Sheetname = str("Sheet" + str(j))
        Sheets(Sheetname).Select()
        for i in range(1, pagenumb):
            wb = openpyxl.load_workbook(str(str(BOM_output_path) + "catia_bom.xlsx"))
            kss = {"What": "Product Description", "After": "ActiveCell", "LookIn": "xlFormulas", "LookAt": "xlPart",
                   "SearchOrder": "xlByRows", "SearchDirection": "xlNext", "MatchCase": False, "MatchByte": False,
                   "SearchFormat": False}
            Cells.Find(kss)
            ss = kss.Column  # 列數
            sc = kss.Row + i - 1  # 行動
            Cells(sc, ss).Select()  # 位置
            Cells(sc, ss).Offset(RowOffset=1).Select()
            Selection.Copy()
            wb = openpyxl.load_workbook(str(str(BOM_output_path) + "BOM_空白頁.xlsx"))
            kss1 = {"What": "規格", "After": "ActiveCell", "LookIn": "xlFormulas", "LookAt": "xlPart",
                    "SearchOrder": "xlByRows", "SearchDirection": "xlNext", "MatchCase": False, "MatchByte": False,
                    "SearchFormat": False}
            Cells.Find(kss1)
            ss1 = kss.Column  # 列數
            sc1 = kss.Row + i - 1  # 行動
            Cells(sc1, ss1).Select()  # 位置
            Cells(sc1, ss1).Offset(RowOffset=1).Select()
            ActiveSheet.Paste()
            loops += 1
        page0 -= 1
        if page0 == 0:
            pagenumb = cunt - 30 * page


def decide_Pa(cunt, page):
    loops = 0
    if page < 1:
        pagenumb = cunt
    if page >= 1:
        pagenumb = 30
    page0 = page
    for j in range(1, page + 1):
        Sheetname = str("Sheet" + str(j))
        Sheets(Sheetname).Select()
        for i in range(1, pagenumb):
            wb = openpyxl.load_workbook(str(str(BOM_output_path) + "catia_bom.xlsx"))
            kss = {"What": "Page", "After": "ActiveCell", "LookIn": "xlFormulas", "LookAt": "xlPart",
                   "SearchOrder": "xlByRows", "SearchDirection": "xlNext", "MatchCase": False, "MatchByte": False,
                   "SearchFormat": False}
            Cells.Find(kss)
            ss = kss.Column  # 列數
            sc = kss.Row + i - 1  # 行動
            Cells(sc, ss).Select()  # 位置
            Cells(sc, ss).Offset(RowOffset=1).Select()
            Selection.Copy()
            wb = openpyxl.load_workbook(str(str(BOM_output_path) + "BOM_空白頁.xlsx"))
            kss1 = {"What": "頁碼", "After": "ActiveCell", "LookIn": "xlFormulas", "LookAt": "xlPart",
                    "SearchOrder": "xlByRows", "SearchDirection": "xlNext", "MatchCase": False, "MatchByte": False,
                    "SearchFormat": False}
            Cells.Find(kss1)
            ss1 = kss.Column  # 列數
            sc1 = kss.Row + i - 1  # 行動
            Cells(sc1, ss1).Select()  # 位置
            Cells(sc1, ss1).Offset(RowOffset=1).Select()
            ActiveSheet.Paste()
            loops += 1
        page0 -= 1
        if page0 == 0:
            pagenumb = cunt - 30 * page


def decide_cost():
    CB_cost()
    MS_cost()
    MSB_cost()
    MSP_cost()
    LDZB_cost()


def draw_block(cunt, page):
    for CBi in range(1, page + 1):
        Sheetname = str("Sheet" + str(CBi))
        Sheets(Sheetname).Select()
        ActiveWindow.SmallScroll(Down=21)
        Range("B7:G36").Select()
        ActiveWindow.SmallScroll(Down=-12)
        Selection.Borders(xlDiagonalDown).LineStyle = xlNone
        Selection.Borders(xlDiagonalUp).LineStyle = xlNone

        Selection.Borders(xlEdgeLeft)
        LineStyle = xlContinuous
        Weight = xlThin
        ColorIndex = xlAutomatic

        Selection.Borders(xlEdgeTop)
        LineStyle = xlContinuous
        Weight = xlThin
        ColorIndex = xlAutomatic

        Selection.Borders(xlEdgeBottom)
        LineStyle = xlContinuous
        Weight = xlThin
        ColorIndex = xlAutomatic

        Selection.Borders(xlEdgeRight)
        LineStyle = xlContinuous
        Weight = xlThin
        ColorIndex = xlAutomatic

        Selection.Borders(xlInsideVertical)
        LineStyle = xlContinuous
        Weight = xlThin
        ColorIndex = xlAutomatic

        Selection.Borders(xlInsideHorizontal)
        LineStyle = xlContinuous
        Weight = xlThin
        ColorIndex = xlAutomatic

        # ------置中
        # Range("A7:H36").Select()
        # Selection("A7:H36")
        # HorizontalAlignment = xlCenter
        # VerticalAlignment = xlCenter
        # WrapText = False
        # Orientation = 0
        # AddIndent = False
        # IndentLevel = 0
        # ShrinkToFit = False
        # ReadingOrder = xlContext
        # MergeCells = False


def CB_cost():
    CB3_5 = 1
    CB3_6 = 1
    CB3_8 = 1
    CB3_10 = 1
    CB3_12 = 1
    CB3_14 = 1
    CB3_15 = 1
    CB3_16 = 1
    CB3_18 = 1
    CB3_20 = 1
    CB3_22 = 2
    CB3_25 = 2
    CB3_30 = 3
    CB3_35 = 3
    CB3_40 = 3
    CB3_45 = 5
    CB3_50 = 7
    CB3_55 = 7
    CB3_60 = 22
    CB3_65 = 23
    CB4_5 = 1
    CB4_6 = 1
    CB4_8 = 1
    CB4_10 = 1
    CB4_12 = 1
    CB4_14 = 1
    CB4_15 = 1
    CB4_16 = 1
    CB4_18 = 1
    CB4_20 = 1
    CB4_22 = 1
    CB4_25 = 1
    CB4_30 = 2
    CB4_35 = 2
    CB4_40 = 3
    CB4_45 = 3
    CB4_50 = 3
    CB4_55 = 6
    CB4_60 = 7
    CB4_65 = 14
    CB4_70 = 16
    CB4_75 = 18
    CB5_5 = 4
    CB5_6 = 4
    CB5_8 = 1
    CB5_10 = 1
    CB5_12 = 1
    CB5_14 = 1
    CB5_15 = 1
    CB5_16 = 1
    CB5_18 = 1
    CB5_20 = 1
    CB5_22 = 1
    CB5_25 = 1
    CB5_30 = 2
    CB5_35 = 2
    CB5_40 = 3
    CB5_45 = 3
    CB5_50 = 3
    CB5_55 = 4
    CB5_60 = 4
    CB5_65 = 5
    CB5_70 = 5
    CB5_75 = 6
    CB5_80 = 7
    CB5_85 = 11
    CB5_90 = 11
    CB5_95 = 18
    CB5_100 = 18
    CB5_110 = 23
    CB5_120 = 27
    CB5_130 = 31
    CB5_140 = 34
    CB6_6 = 5
    CB6_8 = 2
    CB6_10 = 1
    CB6_12 = 1
    CB6_14 = 1
    CB6_15 = 1
    CB6_16 = 1
    CB6_18 = 1
    CB6_20 = 2
    CB6_22 = 2
    CB6_25 = 2
    CB6_30 = 2
    CB6_35 = 3
    CB6_40 = 3
    CB6_45 = 3
    CB6_50 = 4
    CB6_55 = 4
    CB6_60 = 4
    CB6_65 = 5
    CB6_70 = 5
    CB6_75 = 6
    CB6_80 = 7
    CB6_85 = 7
    CB6_90 = 8
    CB6_95 = 9
    CB6_100 = 10
    CB6_110 = 14
    CB6_120 = 29
    CB6_130 = 36
    CB6_140 = 47
    CB6_150 = 53
    CB8_8 = 5
    CB8_10 = 3
    CB8_12 = 3
    CB8_14 = 3
    CB8_15 = 3
    CB8_16 = 3
    CB8_18 = 3
    CB8_20 = 3
    CB8_22 = 3
    CB8_25 = 3
    CB8_30 = 3
    CB8_35 = 3
    CB8_40 = 4
    CB8_45 = 4
    CB8_50 = 4
    CB8_55 = 5
    CB8_60 = 6
    CB8_65 = 6
    CB8_70 = 7
    CB8_75 = 8
    CB8_80 = 9
    CB8_85 = 11
    CB8_90 = 12
    CB8_95 = 13
    CB8_100 = 13
    CB8_110 = 16
    CB8_120 = 18
    CB8_130 = 26
    CB8_140 = 28
    CB8_150 = 32
    CB8_160 = 43
    CB8_200 = 88
    CB10_10 = 6
    CB10_12 = 6
    CB10_15 = 5
    CB10_20 = 4
    CB10_25 = 4
    CB10_30 = 5
    CB10_35 = 5
    CB10_40 = 5
    CB10_45 = 6
    CB10_50 = 6
    CB10_55 = 7
    CB10_60 = 8
    CB10_65 = 9
    CB10_70 = 9
    CB10_75 = 11
    CB10_80 = 12
    CB10_85 = 13
    CB10_90 = 14
    CB10_95 = 15
    CB10_100 = 15
    CB10_110 = 18
    CB10_120 = 22
    CB10_130 = 26
    CB10_140 = 28
    CB10_150 = 32
    CB10_160 = 55
    CB10_170 = 63
    CB10_180 = 69
    CB10_190 = 81
    CB10_200 = 86
    CB10_210 = 102
    CB12_15 = 9
    CB12_20 = 7
    CB12_25 = 7
    CB12_30 = 8
    CB12_35 = 8
    CB12_40 = 8
    CB12_45 = 9
    CB12_50 = 10
    CB12_55 = 11
    CB12_60 = 12
    CB12_65 = 13
    CB12_70 = 14
    CB12_75 = 14
    CB12_80 = 16
    CB12_85 = 19
    CB12_90 = 21
    CB12_95 = 21
    CB12_100 = 21
    CB12_110 = 25
    CB12_120 = 27
    CB12_130 = 31
    CB12_140 = 35
    CB12_150 = 37
    CB12_160 = 44
    CB12_170 = 48
    CB12_180 = 52
    CB12_190 = 55
    CB12_200 = 60
    CB12_210 = 68
    CB12_220 = 74
    CB12_230 = 80
    CB12_240 = 85
    CB12_250 = 90
    CB12_260 = 96
    CB12_270 = 256
    CB12_280 = 274
    CB12_290 = 292
    for CBi in range(1, page + 1):
        Sheetname = str("Sheet" + str(CBi))
        Sheets(Sheetname).Select()
        CBC0 = None
        for CBj in range(1, 30):
            CBK = {"What": "CB", "After": "ActiveCell", "LookIn": "xlFormulas", "LookAt": "xlPart",
                   "SearchOrder": "xlByRows",
                   "SearchDirection": "xlNext", "MatchCase": False, "MatchByte": False, "SearchFormat": False}
            Cells.Find(CBK)
            if CBK == None:
                break
            CBk1 = CBK.Row
            CBk2 = CBK.Column
            Cells(CBk1, CBk2).Select()
            CBA = ActiveCell.FormulaR1C1
            CBC_CBj = CBA
            if CBC_CBj == CBC_0:
                break
            if CBj == 1:
                CBC_0 = CBA
            if inStr(CBA, "CB") == 1:
                CBV = 2
            if inStr(CBA, "MS") == 1:
                CBV = 2
            if inStr(CBA, "MSP") == 1:
                CBV = 3
            if inStr(CBA, "MSB") == 1:
                CBV = 3
            if inStr(CBA, "LDZB") == 1:
                CBV = 4
            CBv1 = InStr(CBA, "-")
            CBv2 = Mid(CBA, CBV + 1, CBv1 - CBV - 1)
            CBv3 = Mid(CBA, CBv1 + 1)
            CBv4 = CDbl(CBv3)
            if CBv2 == 3:
                CBcost = CB3_(CBv4)
            if CBv2 == 4:
                CBcost = CB4_(CBv4)
            if CBv2 == 5:
                CBcost = CB5_(CBv4)
            if CBv2 == 6:
                CBcost = CB6_(CBv4)
            if CBv2 == 8:
                CBcost = CB8_(CBv4)
            if CBv2 == 10:
                CBcost = CB10_(CBv4)
            if CBv2 == 12:
                CBcost = CB12_(CBv4)

            # 數量
            Cells(CBk1, CBk2 + 1).Select()
            CBb = ActiveCell.FormulaR1C1

            co = {"What": "金額", "After": "ActiveCell", "LookIn": "xlFormulas", "LookAt": "xlPart",
                  "SearchOrder": "xlByRows",
                  "SearchDirection": "xlNext", "MatchCase": False, "MatchByte": False, "SearchFormat": False}
            Cells.Find(co)
            co1 = co.Column
            Cells(CBk1, co1).Select()
            ActiveCell.FormulaR1C1 = CBcost * CBb


def MS_cost():
    MS1_6 = 10
    MS1_8 = 10
    MS1_10 = 10
    MS2_6 = 7
    MS2_8 = 7
    MS2_10 = 5
    MS2_15 = 5
    MS2_20 = 7
    MS3_6 = 7
    MS3_8 = 7
    MS3_10 = 3
    MS3_15 = 3
    MS3_20 = 3
    MS3_25 = 7
    MS3_30 = 7
    MS3_35 = 7
    MS3_40 = 7
    MS4_8 = 7
    MS4_10 = 4
    MS4_15 = 4
    MS4_20 = 4
    MS4_25 = 7
    MS4_30 = 7
    MS4_35 = 7
    MS4_40 = 7
    MS4_45 = 7
    MS4_50 = 7
    MS5_8 = 7
    MS5_10 = 5
    MS5_15 = 5
    MS5_20 = 5
    MS5_25 = 5
    MS5_30 = 5
    MS5_35 = 7
    MS5_40 = 7
    MS5_45 = 7
    MS5_50 = 7
    MS6_8 = 10
    MS6_10 = 10
    MS6_15 = 6
    MS6_20 = 6
    MS6_25 = 6
    MS6_30 = 6
    MS6_35 = 10
    MS6_40 = 10
    MS6_45 = 10
    MS6_50 = 10
    MS6_55 = 10
    MS6_60 = 10
    MS8_10 = 14
    MS8_15 = 14
    MS8_20 = 14
    MS8_25 = 14
    MS8_30 = 10
    MS8_35 = 14
    MS8_40 = 10
    MS8_45 = 14
    MS8_50 = 10
    MS8_55 = 14
    MS8_60 = 14
    MS8_65 = 14
    MS8_70 = 14
    MS8_80 = 14
    MS10_15 = 24
    MS10_20 = 24
    MS10_25 = 24
    MS10_30 = 24
    MS10_35 = 24
    MS10_40 = 14
    MS10_45 = 24
    MS10_50 = 14
    MS10_55 = 24
    MS10_60 = 14
    MS10_65 = 24
    MS10_70 = 24
    MS10_80 = 24
    MS12_20 = 25
    MS12_25 = 25
    MS12_30 = 25
    MS12_35 = 25
    MS12_40 = 20
    MS12_45 = 25
    MS12_50 = 20
    MS12_55 = 25
    MS12_60 = 20
    MS12_70 = 25
    MS12_80 = 25
    for CBi in range(1, page + 1):
        Sheetname = str("Sheet" + str(CBi))
        Sheets(Sheetname).Select()
        CBC_0 = None
        for CBj in range(1, 30):
            CBK = {"What": "MS", "After": "ActiveCell", "LookIn": "xlFormulas", "LookAt": "xlPart",
                   "SearchOrder": "xlByRows",
                   "SearchDirection": "xlNext", "MatchCase": False, "MatchByte": False, "SearchFormat": False}
            Cells.Find(CBK)
            if CBK == None:
                break
            CBk1 = CBK.Row
            CBk2 = CBK.Column
            Cells(CBk1, CBk2).Select()
            CBA = ActiveCell.FormulaR1C1
            CBC_CBj = CBA
            if CBC_CBj == CBC_0:
                break
            if CBj == 1:
                CBC_0 = CBA
            if inStr(CBA, "CB") == 1:
                CBV = 2
            if inStr(CBA, "MS") == 1:
                CBV = 2
            if inStr(CBA, "MSP") == 1:
                CBV = 3
            if inStr(CBA, "MSB") == 1:
                CBV = 3
            if inStr(CBA, "LDZB") == 1:
                CBV = 4
            CBv1 = InStr(CBA, "-")
            CBv2 = Mid(CBA, CBV + 1, CBv1 - CBV - 1)
            CBv3 = Mid(CBA, CBv1 + 1)
            CBv4 = CDbl(CBv3)
            if CBV == 2:
                CBcost = None
            if CBv2 == 1:
                CBcost = MS1_(CBv4)
            if CBv2 == 2:
                CBcost = MS2_(CBv4)
            if CBv2 == 3:
                CBcost = MS3_(CBv4)
            if CBv2 == 4:
                CBcost = MS4_(CBv4)
            if CBv2 == 5:
                CBcost = MS5_(CBv4)
            if CBv2 == 6:
                CBcost = MS6_(CBv4)
            if CBv2 == 8:
                CBcost = MS8_(CBv4)
            if CBv2 == 10:
                CBcost = MS10_(CBv4)
            if CBv2 == 12:
                CBcost = MS12_(CBv4)
            else:
                break  # 不確定
            Cells(CBk1, CBk2 + 1).Select()
            CBb = ActiveCell.FormulaR1C1
            co = {"What": "金額", "After": "ActiveCell", "LookIn": "xlFormulas", "LookAt": "xlPart",
                  "SearchOrder": "xlByRows",
                  "SearchDirection": "xlNext", "MatchCase": False, "MatchByte": False, "SearchFormat": False}
            Cells.Find(co)
            co1 = co.Column
            Cells(CBk1, co1).Select()
            ActiveCell.FormulaR1C1 = CBcost * CBb


def MSB_cost():
    MSB4_10 = 25
    MSB4_15 = 20
    MSB4_20 = 20
    MSB4_25 = 20
    MSB4_30 = 20
    MSB4_35 = 20
    MSB4_40 = 20
    MSB4_45 = 25
    MSB5_10 = 25
    MSB5_15 = 20
    MSB5_20 = 20
    MSB5_25 = 20
    MSB5_30 = 20
    MSB5_35 = 20
    MSB5_40 = 20
    MSB5_45 = 25
    MSB5_50 = 25
    MSB6_10 = 34
    MSB6_15 = 27
    MSB6_20 = 27
    MSB6_25 = 27
    MSB6_30 = 27
    MSB6_35 = 27
    MSB6_40 = 27
    MSB6_45 = 34
    MSB6_50 = 34
    MSB6_55 = 34
    MSB6_60 = 34
    MSB6_65 = 34
    MSB6_70 = 34
    MSB8_10 = 29
    MSB8_15 = 23
    MSB8_20 = 23
    MSB8_25 = 23
    MSB8_30 = 23
    MSB8_35 = 23
    MSB8_40 = 23
    MSB8_45 = 23
    MSB8_50 = 29
    MSB8_55 = 29
    MSB8_60 = 29
    MSB8_65 = 29
    MSB8_70 = 29
    MSB8_75 = 29
    MSB8_80 = 29
    MSB8_85 = 29
    MSB8_90 = 29
    MSB8_95 = 36
    MSB8_100 = 36
    MSB8_105 = 36
    MSB8_110 = 36
    MSB8_115 = 36
    MSB8_120 = 36
    MSB10_10 = 43
    MSB10_15 = 31
    MSB10_20 = 31
    MSB10_25 = 31
    MSB10_30 = 31
    MSB10_35 = 31
    MSB10_40 = 31
    MSB10_45 = 31
    MSB10_50 = 31
    MSB10_55 = 31
    MSB10_60 = 43
    MSB10_65 = 43
    MSB10_70 = 43
    MSB10_75 = 43
    MSB10_80 = 43
    MSB10_85 = 43
    MSB10_90 = 43
    MSB10_95 = 43
    MSB10_100 = 43
    MSB10_105 = 43
    MSB10_110 = 43
    MSB10_115 = 43
    MSB10_120 = 43
    MSB12_10 = 65
    MSB12_15 = 31
    MSB12_20 = 31
    MSB12_25 = 31
    MSB12_30 = 31
    MSB12_35 = 31
    MSB12_40 = 31
    MSB12_45 = 31
    MSB12_50 = 31
    MSB12_55 = 31
    MSB12_60 = 50
    MSB12_65 = 50
    MSB12_70 = 50
    MSB12_75 = 50
    MSB12_80 = 65
    MSB12_85 = 65
    MSB12_90 = 65
    MSB12_95 = 65
    MSB12_100 = 65
    for CBi in range(1, page + 1):
        Sheetname = str("Sheet" + str(CBi))
        Sheets(Sheetname).Select()
        CBC_0 = None
        for CBj in range(1, 30):
            CBK = {"What": "MSB", "After": "ActiveCell", "LookIn": "xlFormulas", "LookAt": "xlPart",
                   "SearchOrder": "xlByRows",
                   "SearchDirection": "xlNext", "MatchCase": False, "MatchByte": False, "SearchFormat": False}
            Cells.Find(CBK)
            if CBK == None:
                break
            CBk1 = CBK.Row
            CBk2 = CBK.Column
            Cells(CBk1, CBk2).Select()
            CBA = ActiveCell.FormulaR1C1
            CBC_CBj = CBA
            if CBC_CBj == CBC_0:
                break
            if CBj == 1:
                CBC_0 = CBA
            if inStr(CBA, "CB") == 1:
                CBV = 2
            if inStr(CBA, "MS") == 1:
                CBV = 2
            if inStr(CBA, "MSP") == 1:
                CBV = 3
            if inStr(CBA, "MSB") == 1:
                CBV = 3
            if inStr(CBA, "LDZB") == 1:
                CBV = 4
            CBv1 = InStr(CBA, "-")
            CBv2 = Mid(CBA, CBV + 1, CBv1 - CBV - 1)
            CBv3 = Mid(CBA, CBv1 + 1)
            CBv4 = CDbl(CBv3)
            if CBv2 == 4:
                CBcost = MS4_(CBv4)
            if CBv2 == 5:
                CBcost = MS5_(CBv4)
            if CBv2 == 6:
                CBcost = MS6_(CBv4)
            if CBv2 == 8:
                CBcost = MS8_(CBv4)
            if CBv2 == 10:
                CBcost = MS10_(CBv4)
            if CBv2 == 12:
                CBcost = MS12_(CBv4)
            else:
                break  # 不確定
            # 數量
            Cells(CBk1, CBk2 + 1).Select()
            CBb = ActiveCell.FormulaR1C1
            co = {What: "金額", After: ActiveCell, LookIn: xlFormulas, LookAt: xlPart, SearchOrder: xlByRows,
                  SearchDirection: xlNext, MatchCase: False, MatchByte: False, SearchFormat: False}
            Cells.Find(co)
            co1 = co.Column
            Cells(CBk1, co1).Select()
            ActiveCell.FormulaR1C1 = CBcost * CBb


def MSP_cost():
    MSP32_190 = 262
    for CBi in range(1, page + 1):
        Sheetname = str("Sheet" + str(CBi))
        Sheets(Sheetname).Select()
        CBC_0 = None
        for CBj in range(1, 30):
            CBK = {What: "MSP", After: ActiveCell, LookIn: xlFormulas, LookAt: xlPart, SearchOrder: xlByRows,
                   SearchDirection: xlNext, MatchCase: False, MatchByte: False, SearchFormat: False}
            Cells.Find(CBK)
            if CBK == None:
                break
            CBk1 = CBK.Row
            CBk2 = CBK.Column
            Cells(CBk1, CBk2).Select()
            CBA = ActiveCell.FormulaR1C1
            CBC_CBj = CBA
            if CBC_CBj == CBC_0:
                break
            if CBj == 1:
                CBC_0 = CBA
            if inStr(CBA, "CB") == 1:
                CBV = 2
            if inStr(CBA, "MS") == 1:
                CBV = 2
            if inStr(CBA, "MSP") == 1:
                CBV = 3
            if inStr(CBA, "MSB") == 1:
                CBV = 3
            if inStr(CBA, "LDZB") == 1:
                CBV = 4
            CBv1 = inStr(CBA, "-")
            CBv2 = Mid(CBA, CBV + 1, CBv1 - CBV - 1)
            CBv3 = Mid(CBA, CBv1 + 1)
            CBv4 = CDbl(CBv3)
            if CBv2 == 32:
                CBcost = MS32_(CBv4)
            else:
                break  # 不確定
            # 數量
            Cells(CBk1, CBk2 + 1).Select()
            CBb = ActiveCell.FormulaR1C1
            co = {What: "金額", After: ActiveCell, LookIn: xlFormulas, LookAt: xlPart, SearchOrder: xlByRows,
                  SearchDirection: xlNext, MatchCase: False, MatchByte: False, SearchFormat: False}
            Cells.Find(co)
            co1 = co.Column
            Cells(CBk1, co1).Select()
            ActiveCell.FormulaR1C1 = CBcost * CBb


def LDZB_cost():
    LDBZ32_80 = 1325
    for CBi in range(1, page + 1):
        Sheetname = str("Sheet" + str(CBi))
        Sheets(Sheetname).Select()
        CBC_0 = None
        for CBj in range(1, 30):
            CBK = {What: "LDBZ", After: ActiveCell, LookIn: xlFormulas, LookAt: xlPart, SearchOrder: xlByRows,
                   SearchDirection: xlNext, MatchCase: False, MatchByte: False, SearchFormat: False}
            Cells.Find(CBK)
            if CBK == None:
                break
            CBk1 = CBK.Row
            CBk2 = CBK.Column
            Cells(CBk1, CBk2).Select()
            CBA = ActiveCell.FormulaR1C1
            CBC_CBj = CBA
            if CBC_CBj == CBC_0:
                break
            if CBj == 1:
                CBC_0 = CBA
            if inStr(CBA, "CB") == 1:
                CBV = 2
            if inStr(CBA, "MS") == 1:
                CBV = 2
            if inStr(CBA, "MSP") == 1:
                CBV = 3
            if inStr(CBA, "MSB") == 1:
                CBV = 3
            if inStr(CBA, "LDZB") == 1:
                CBV = 4
            CBv1 = inStr(CBA, "-")
            CBv2 = Mid(CBA, CBV + 1, CBv1 - CBV - 1)
            CBv3 = Mid(CBA, CBv1 + 1)
            CBv4 = CDbl(CBv3)
            if CBv2 == 32:
                CBcost = MS32_(CBv4)
            else:
                break  # 不確定
            # 數量
            Cells(CBk1, CBk2 + 1).Select()
            CBb = ActiveCell.FormulaR1C1
            co = {What: "金額", After: ActiveCell, LookIn: xlFormulas, LookAt: xlPart, SearchOrder: xlByRows,
                  SearchDirection: xlNext, MatchCase: False, MatchByte: False, SearchFormat: False}
            Cells.Find(co)
            co1 = co.Column
            Cells(CBk1, co1).Select()
            ActiveCell.FormulaR1C1 = CBcost * CBb


def Adjustment():
    for i in range(1, page + 1):
        Sheetname = str(str("Sheet") + str(i))
        Sheets(Sheetname).Select()
        # --------------------------------------------------------調整欄寬至適當大小
        ActiveSheet.Range("B7").Select()
        Range(Selection, Selection.End(xlDown)).Select()
        ActiveWindow.RangeSelection()
        Columns.AutoFit()
        Rows.AutoFit()
        # --------------------------------------------------------調整欄寬至適當大小

        # --------------------------------------------------------文字置中
        Worksheets(Sheetname).Activate()
        ActiveSheet.UsedRange.Select()
        Range("A1:H37").Borders.LineStyle = xlContinuous  # EXCEL範圍內加上框線
        Selection()
        HorizontalAlignment = xlCenter
        VerticalAlignment = xlCenter
        WrapText = False
        Orientation = 0
        AddIndent = False
        IndentLevel = 0
        ShrinkToFit = False
        ReadingOrder = xlContext
        # --------------------------------------------------------文字置中
        Selection.Font.Name = "標楷體"


BOMMaking()
